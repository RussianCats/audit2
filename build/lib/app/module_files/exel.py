from app import *
from app.services.logging import logger
from app.forms import *
import os

import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Alignment
from copy import deepcopy
from pathlib import Path

def styleTable(table, font_size):
    # словарь с размерами столбцов
    cols_dict = {}

    # проходимся по всем строкам документа
    for row in table.rows:
        # теперь по ячейкам каждой строки
        for cell in row:
            # получаем букву текущего столбца
            letter = cell.column_letter
            # если в ячейке записаны данные
            if cell.value:
                # устанавливаем в ячейке размер шрифта
                cell.font = Font(name='Cambria', size=font_size)
                # если в значении ячейки есть переносы строк, разделяем текст
                # и находим максимальную длину среди всех строк
                lines = str(cell.value).split('\n')
                len_cell = max(len(line) for line in lines)
                # обновляем максимальную длину, если текущая больше сохраненной
                len_cell_dict = cols_dict.get(letter, 0)
                if len_cell > len_cell_dict:
                    cols_dict[letter] = len_cell
                    # расчет новой ширины колонки (примерный расчет)
                    new_width_col = (len_cell + 2) * font_size * 0.1
                    # применение новой ширины столбца
                    table.column_dimensions[letter].width = new_width_col + 1

    return table


# перенос строк в тексте
def tableWrapText(table):
    for row in table.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    return table


def readExcelFile(_path, _size):
    #  модель данных 
    # - Адрес	
    # - Структурное подразделение (отделение)	
    # - Кабинет №	
    # - ФИО сотрудника
    # - Сетевое имя компьютера/сервера	
    # - Инвентарный № системного блока	
    # - Монитор (модель)	
    # - Клавиатура (модель)	
    # - Мышь (модель)

    listInfoReport = []
    try:
        workbook = load_workbook(filename=_path, read_only=True)

        # Итерация по всем листам в книге
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Пропускаем первую строку (заголовки)
            for row in sheet.iter_rows(min_row=2, values_only=True, max_col=_size):
                
                # Замена пустых ячеек на пустую строку и сохранение строки в список
                row_data = [cell if cell is not None else "" for cell in row]
                
                
                if not all(cell == "" for cell in row_data):
                    report = deepcopy(infoReport) 

                    pattern = re.compile(r"^([a-zA-Z][-a-zA-Z0-9.]*)(?:_\[([a-zA-Z0-9-]+)\])?$")
                    match = pattern.match(row_data[4])
                    if match:
                        report["hostname"] = match.group(1)
                        report["uuid"] = match.group(2) if match.group(2) else None
                    else:
                        report["hostname"] = row_data[4]
                    report["hostname"] =  report["hostname"].strip() 


                        # logger.warn(f"Не правильный шаблон сетевогоимени в таблице {_path} ")

                    
                    report["org"]["address"] = row_data[0]
                    report["org"]["structural"] = re.split(r'[,;]\s*|\r?\n', row_data[1])
                    report["org"]["cabinet"] = row_data[2]
                    report["org"]["FIO"] = re.split(r'[,;]\s*|\r?\n', row_data[3])
                    report["org"]["inventory"] = row_data[5]
                    report["monitor"] = re.split(r'[,;]\s*|\r?\n', row_data[6])
                    report["keyboard"] = row_data[7]
                    report["mouse"] = row_data[8]


                    listInfoReport.append(report)

        
        workbook.close()
    except FileNotFoundError:
        print(f"Файл не найден: {_path}")
    except InvalidFileException:
        print(f"Невозможно открыть файл: {_path}. Проверьте, является ли файл корректным Excel-файлом.")
    except Exception as e:
        print(f"Произошла ошибка при чтении файла {_path}: {e}")

    return listInfoReport


def writeExelTable(_listReports):
    

    getTemplatesAdd = ["Вид (Сервер/АРМ/Иное техническое средство) (обязательное поле)", "Сетевое имя (обязательное поле для АРМ и сервера)", \
        "Тип технического средства (обязательное поле для иного технического средства)", "Модель (обязательное поле для иного технического средства)", \
        "Условное наименование (обязательное поле)", "Заводской/ инвентарный номер", "IP-адрес", "MAC-адрес" ]
    
    # getTemplatesUpdatePlatform = ["Техническое средство (обязательное поле)", "Идентификационная информация (обязательное поле)", "Система (сеть)", \
    #     "Программные комплексы", "Тип", "Роль сервера", "Группы технических средств", "Видеоадаптер", "Приводы CD/DVD", "Системная память", "Материнская плата", \
    #     "Дисковые устройства", "Клавиатура", "Мышь", "Процессор", "Монитор", "Вид микропрограммного обеспечения (обязательное поле)", "Наименование программного обеспечения", \
    #     "Версия микропрограммного обеспечения", "Вид системного программного обеспечения", "Наименование системного программного обеспечения", \
    #     "Версия системного программного обеспечения", "Программная оболочка", "Драйвер", "Утилита", "Загрузчик операционной системы", "Помещение / Адрес поставщика вычислительных услуг (обязательное поле)", \
    #     "Структурное подразделение", "Доступ сотрудников" ]
    
    getTemplatesUpdate = ["Техническое средство (обязательное поле)", "Идентификационная информация (обязательное поле)", "Объект информатизации", \
        "Программное обеспечение [Версия ПО]", "Тип", "Роль сервера", "Группы технических средств", "Процессор", "Материнская плата", "Системная память", "Дисковые устройства", \
        "Видеоадаптер", "Приводы CD/DVD", "Монитор", "Клавиатура", "Мышь", "Помещение / Поставщик вычислительных услуг", \
        "Адрес помещения / поставщика вычислительных услуг", "Доступ сотрудников" ]
        
    # getTemplatesOAIS = ["Сетевое имя", "Доступ сотрудников", "Структурные подразделения", "СЗИ (Версия/Дата установки)", "Приложения и версия"]
    
    getTemplatesDIR = ["Сетевое имя", "Кабинет", "Сотрудники", "Инвентарный номер", "IP", "MAC", "OC + билд", "СЗИ (Версия/Дата установки)"]

    getTemplatesCabinet = ["Наименование помещения (обязательное поле)", "Адрес помещения (обязательное поле)", "Структурные подразделения", "Сотрудники, имеющие доступ к помещению"]

    # системное програмнное обеспечение
    getTemplatesSoftware = ["Наименование ПО (обязательное)", "Тип ПО (обязательное)", "Категория ПО", "Версия", "Разработчик", "Адрес разработчика"]

    # для создания таблицы кабинетов
    _arr_cabinet = {}
    # список ОС
    _arr_os = []

    
    # Создаем новую книгу Excel
    wbUpdate = Workbook()
    wbAdd = Workbook()
    wbDIR = Workbook()
    wbCabinet = Workbook()
    wbSoftware = Workbook()

    # Получаем активный лист
    _table_Update = wbUpdate.active
    _table_Add = wbAdd.active
    _table_DIR = wbDIR.active
    _table_Cabinet = wbCabinet.active
    _table_Software = wbSoftware.active

    # добавить шаблоны
    _table_Update.append(getTemplatesUpdate)
    _table_Add.append(getTemplatesAdd)
    _table_DIR.append(getTemplatesDIR)
    _table_Cabinet.append(getTemplatesCabinet)
    _table_Software.append(getTemplatesSoftware)

    # Заполняем лист данными
    for report in _listReports:

        # для учетов кабинетов
        key = (report['org']['cabinet'], report['org']['address'])
        if key not in _arr_cabinet:
            _arr_cabinet[key] = {'structural': set(), 'FIO': set()}
        _arr_cabinet[key]['structural'].update(report['org']['structural'])
        _arr_cabinet[key]['FIO'].update(report['org']['FIO'])

        # список ОС
        _arr_os.append(report["os"]["small"])
        
        tmp_ip = ""
        tmp_mac = ""
        # таблица добавить из файла
        try:
            tmp_ip = f"{report["ip"][0] if (report["duplicate"] == True) or config.CONFIG.ADD_IP_TABLE else ""}"
        except:
            logger.warn(f"Не удалось найти IP адрес {report["hostname"]}")
            tmp_ip = ""
        try:
            tmp_mac =  f"{report["mac"][0] if (report["duplicate"] == True) or config.CONFIG.ADD_MAC_TABLE else ""}"
        except:
            logger.warn(f"Не удалось найти MAC адрес {report["hostname"]}")
            tmp_mac = ""

        _table_Add.append([
            "АРМ",
            f"{report["hostname"]}",
            "",
            "",
            "АРМ",
            f"{report["org"]["inventory"]}",
            tmp_ip,
            tmp_mac
 
        ])

        # таблица обновить из файла
        _table_Update.append([
            "АРМ",
            f"Сетевое имя: \"{report["hostname"]}\"; \n\
Заводской/ инвентарный номер: \"{"б/н" if report["org"]["inventory"] == "" else report["org"]["inventory"]}\"; \n\
IP-адрес: \"{tmp_ip}\"; \n\
MAC-адрес: \"{tmp_mac}\"; \n\
Модель: \"\"",
            "",
            f'{report["os"]["small"]}',
            "Физический",
            "",
            "",
            f"{report["cpu"]}",
            f"{report["motherboard"]["vendor"]} {report["motherboard"]["model"]}",
            f"{report["ram"]}",
            f"{report["disk"]}",
            f"{report["videoadapter"]}",
            f"{report["cd_dvd"]}",
            f"{'; '.join(report['monitor'])}",
            f"{report["keyboard"]}",
            f"{report["mouse"]}",
            f"{report["org"]["cabinet"]}",
            f"{report["org"]["address"]}",
            f"{'; '.join(report['org']['FIO'])}"

        ])

        str_ipt = []
        for ipt in report["ipt_list"]:
            str_ipt.append(f"{ipt[0]} ({ipt[1]} / {ipt[2]})")
        
        # таблица ДИР
        _table_DIR.append([
            f"{report["hostname"]}",
            f"{report["org"]["cabinet"]}",
            f"{'; '.join(report["org"]["FIO"])}",
            f"{report["org"]["inventory"]}",
            f"{';\n'.join(report['ip'])}",
            f"{';\n'.join(report['mac'])}",
            f"{report["os"]["full"]}",
            f"{';\n'.join(str_ipt)}"
        ])

    
    # # таблица системно программного обеспечения
    _arr_os = [i for i in _arr_os if i != ""]
    _arr_os = list(set(_arr_os))
    _arr_os.sort()

    for os_report in _arr_os:

        _table_Software.append([
            f"{os_report}",
            "Системное программное обеспечение",
            "Операционная система"
        ])

    # Заполнение таблицы кабинетов агрегированными данными
    for (cabinet, address), data in _arr_cabinet.items():
        structural = '; '.join(data['structural'])
        fio = '; '.join(data['FIO'])
        row = [cabinet, address, structural, fio]
        _table_Cabinet.append(row)


    # настройка стилей
    _table_Update = tableWrapText(_table_Update)
    _table_Update = styleTable(_table_Update, 11)
    _table_Add = tableWrapText(_table_Add)
    _table_Add = styleTable(_table_Add, 12)
    _table_DIR = tableWrapText(_table_DIR)
    _table_DIR = styleTable(_table_DIR, 11)
    _table_Cabinet = tableWrapText(_table_Cabinet)
    _table_Cabinet = styleTable(_table_Cabinet, 14)
    _table_Software = tableWrapText(_table_Software)
    _table_Software = styleTable(_table_Software, 14)
    
    try:
        dir_path = Path(config.CONFIG.EXECUTPATH) / "report"
        dir_path.mkdir(parents=True, exist_ok=True)
    except OSError as error:
        logger.info(error)
    
    # Сохраняем книгу   
    wbUpdate.save(Path(config.CONFIG.EXECUTPATH) / "report" / "Обновить средства из файла.xlsx")
    wbAdd.save(Path(config.CONFIG.EXECUTPATH) / "report" / "Добавить средства из файла.xlsx")
    wbDIR.save(Path(config.CONFIG.EXECUTPATH) / "report" / "Таблица ДИР.xlsx")
    wbCabinet.save(Path(config.CONFIG.EXECUTPATH) / "report" / "Кабинеты.xlsx")
    wbSoftware.save(Path(config.CONFIG.EXECUTPATH) / "report" / "ПО.xlsx")






